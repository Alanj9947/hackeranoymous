"""
Excel export service with templates and formatting.
"""

from __future__ import annotations

import io
import os
import tempfile
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.logging import get_logger

logger = get_logger(__name__)

# ── Export Templates ─────────────────────────────────────────────

TEMPLATES = {
    "customer_service_summary": {
        "name": "Customer Service Summary",
        "columns": [
            {"header": "Date", "field": "created_at", "width": 18},
            {"header": "Customer Name", "field": "extracted_data.customer.name", "width": 20},
            {"header": "Email", "field": "extracted_data.customer.email", "width": 25},
            {"header": "Phone", "field": "extracted_data.customer.phone", "width": 15},
            {"header": "Issue Category", "field": "extracted_data.issue.category", "width": 18},
            {"header": "Severity", "field": "extracted_data.issue.severity", "width": 12},
            {"header": "Description", "field": "extracted_data.issue.description", "width": 40},
            {"header": "Status", "field": "extracted_data.resolution.status", "width": 12},
            {"header": "Resolution", "field": "extracted_data.resolution.provided", "width": 35},
            {"header": "Next Steps", "field": "extracted_data.resolution.nextSteps", "width": 30},
            {"header": "Sentiment", "field": "extracted_data.sentiment.overall", "width": 12},
            {"header": "Sentiment Score", "field": "extracted_data.sentiment.score", "width": 14},
            {"header": "Duration (s)", "field": "duration_seconds", "width": 12},
            {"header": "Confidence", "field": "confidence_score", "width": 12},
        ],
        "header_bg": "366092",
        "header_fg": "FFFFFF",
        "alt_row_bg": "D9E1F2",
    },
    "sales_call_analysis": {
        "name": "Sales Call Analysis",
        "columns": [
            {"header": "Date", "field": "created_at", "width": 18},
            {"header": "Prospect Name", "field": "extracted_data.customer.name", "width": 20},
            {"header": "Company", "field": "extracted_data.customer.company", "width": 20},
            {"header": "Product Interest", "field": "extracted_data.issue.category", "width": 20},
            {"header": "Budget Range", "field": "extracted_data.customFields.budget", "width": 15},
            {"header": "Timeline", "field": "extracted_data.customFields.timeline", "width": 15},
            {"header": "Decision Maker", "field": "extracted_data.customFields.decisionMaker", "width": 15},
            {"header": "Next Meeting", "field": "extracted_data.actionItems.0.dueDate", "width": 15},
            {"header": "Sentiment", "field": "extracted_data.sentiment.overall", "width": 12},
            {"header": "Probability", "field": "extracted_data.customFields.closeProbability", "width": 12},
        ],
        "header_bg": "2E7D32",
        "header_fg": "FFFFFF",
        "alt_row_bg": "E8F5E9",
    },
    "recruitment_calls": {
        "name": "Recruitment Calls",
        "columns": [
            {"header": "Candidate Name", "field": "extracted_data.customer.name", "width": 20},
            {"header": "Email", "field": "extracted_data.customer.email", "width": 25},
            {"header": "Phone", "field": "extracted_data.customer.phone", "width": 15},
            {"header": "Position", "field": "extracted_data.issue.category", "width": 20},
            {"header": "Experience", "field": "extracted_data.customFields.experience", "width": 15},
            {"header": "Skills", "field": "extracted_data.customFields.skills", "width": 30},
            {"header": "Interview Stage", "field": "extracted_data.resolution.status", "width": 15},
            {"header": "Feedback", "field": "extracted_data.customFields.feedback", "width": 35},
            {"header": "Rating", "field": "extracted_data.sentiment.score", "width": 10},
            {"header": "Next Steps", "field": "extracted_data.resolution.nextSteps", "width": 25},
        ],
        "header_bg": "6A1B9A",
        "header_fg": "FFFFFF",
        "alt_row_bg": "F3E5F5",
    },
}


class ExcelExportService:
    """Service for exporting extracted call data to Excel files."""

    def export_to_excel(
        self,
        data_rows: List[Dict[str, Any]],
        template_name: str = "customer_service_summary",
        filename: Optional[str] = None,
    ) -> str:
        """
        Export data rows to an Excel file.
        Returns the file path of the generated .xlsx file.
        """
        template = TEMPLATES.get(template_name, TEMPLATES["customer_service_summary"])

        wb = Workbook()
        ws = wb.active
        ws.title = template["name"]

        # Styles
        header_font = Font(bold=True, color=template["header_fg"], size=11)
        header_fill = PatternFill(
            start_color=template["header_bg"],
            end_color=template["header_bg"],
            fill_type="solid",
        )
        alt_fill = PatternFill(
            start_color=template["alt_row_bg"],
            end_color=template["alt_row_bg"],
            fill_type="solid",
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        columns = template["columns"]

        # Write headers
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col["header"])
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = col.get("width", 15)

        # Write data rows
        for row_idx, data in enumerate(data_rows, 2):
            for col_idx, col in enumerate(columns, 1):
                value = self._extract_field(data, col["field"])
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        if data_rows:
            last_col = get_column_letter(len(columns))
            ws.auto_filter.ref = f"A1:{last_col}{len(data_rows) + 1}"

        # Save
        if not filename:
            filename = f"call_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        filepath = os.path.join(tempfile.gettempdir(), f"{filename}.xlsx")
        wb.save(filepath)
        logger.info("excel_exported", filepath=filepath, rows=len(data_rows))
        return filepath

    def export_to_csv(
        self,
        data_rows: List[Dict[str, Any]],
        template_name: str = "customer_service_summary",
        filename: Optional[str] = None,
    ) -> str:
        """Export data rows to a CSV file."""
        import csv

        template = TEMPLATES.get(template_name, TEMPLATES["customer_service_summary"])
        columns = template["columns"]

        if not filename:
            filename = f"call_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        filepath = os.path.join(tempfile.gettempdir(), f"{filename}.csv")

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([col["header"] for col in columns])
            for data in data_rows:
                row = []
                for col in columns:
                    value = self._extract_field(data, col["field"])
                    if isinstance(value, list):
                        value = ", ".join(str(v) for v in value)
                    row.append(value or "")
                writer.writerow(row)

        logger.info("csv_exported", filepath=filepath, rows=len(data_rows))
        return filepath

    def _extract_field(self, obj: dict, path: str) -> Any:
        """Extract a nested field value from a dictionary using dot notation."""
        parts = path.split(".")
        current = obj
        for part in parts:
            if current is None:
                return ""
            # Handle array index like "actionItems.0.dueDate"
            if part.isdigit():
                idx = int(part)
                if isinstance(current, list) and idx < len(current):
                    current = current[idx]
                else:
                    return ""
            elif isinstance(current, dict):
                current = current.get(part)
            else:
                return ""
        return current if current is not None else ""
